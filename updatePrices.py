import time, pyodbc, sys, argparse
import MagicModule as mm, numpy as np, datetime as dt
from openpyxl import Workbook, load_workbook
from os import getcwd

# Default variables
accessFilename = "Magic.accdb"
validationFilename = "validations.txt"
now = dt.datetime.now()
timeWait = 0.1
dir = getcwd() + "\\"
encoding = "latin-1"

# Arugment parsing
parser = argparse.ArgumentParser(description="Create a spreadsheet of Magic: The Gathering card prices")
parser.add_argument("-n", "--name", help="Name of Excel file", default="MagicPrices", type=str)
parser.add_argument("-s", "--stop", help="Stops after a certain count, default of None", default=None, type=int)
parser.add_argument("-q", "--sql", help="Override default query search", default="select * from Cards")
parser.add_argument("-c", "--close", help="Close terminal after fininshing", action="store_true", default=False)
parser.add_argument("-p", "--print", help="Print cards as they are found", action="store_true")
parser.add_argument("-v", "--validate", help="Validate cards", action="store_true")
parser.add_argument("-e", "--export", help="Export into Excel file", action="store_true") # double check this even works
parser.add_argument('-E', "--export_only", help="Only export into Excel file", action="store_true") # same with this
parser.add_argument("--comment", help="Add a comment in the logs", default=None)
parser.add_argument("--use_cache", help="Use cached prices", action="store_true")
args = parser.parse_args()

mm.log(f"INFO: Starting program with arguments: {args}")

# Connection to database
try:
	driverStr = r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ='
	cnxn = pyodbc.connect(driverStr + dir + accessFilename + ";")
	cursor = cnxn.cursor()
	cnxn.setdecoding(pyodbc.SQL_CHAR, encoding=encoding)
	cnxn.setdecoding(pyodbc.SQL_WCHAR, encoding=encoding)
	cnxn.setencoding(encoding=encoding)
except:
	mm.log("ERROR: Failure to connect to drivers and opening database", close=True)

# Open excel workbook
if (".xlsx" not in args.name):args.name=f"{args.name}.xlsx"
workbook = None
try: workbook = load_workbook(filename = args.name)
except Exception as e:
	mm.log(f"WARNING: Failed to open workbook {args.name}, creating it instead")
	workbook = Workbook()
	workbook.save(filename = args.name)
sheet = workbook.active
sheet["A1"] = "Card"
sheet["B1"] = "CN"
sheet["C1"] = "Foiling"
sheet["D1"] = "Set"

# Open text files
if (args.validate): validationFile = open(validationFilename, "w")
else: validationFile = None
if (args.export): exportFile = open("export.csv", "w")
else: exportFile = None

# Get all cards
print("Accessing database for cards")
try:
	cursor.execute(args.sql) # SQL here
except Exception as e:
	mm.log("WARNING: Bad query, defaulting to 'SLECT * FROM Cards'", printMsg=True)
	cursor.execute("SELECT * FROM Cards")
rows = cursor.fetchall()
cards = [mm.Card(c[0], c[1], c[2], c[3], quantity=c[4]) for c in rows]
cursor.close()
cnxn.close()
print("Cards collected. Setting up excel file for new entries")

# Get new column for prices
column = 1
while (sheet[f"{mm.numToCol(column)}1"].value != None): column += 1
column = mm.numToCol(column)
sheet[f"{column}1"] = dt.datetime(now.year, now.month, now.day)
sheet[f"{column}1"].number_format = "mm/dd/yyyy;@"

# Export to csv
if (args.export):
	print("Exporting to export.csv")
	for card in cards:
		if (card.foil == "No"): foiling = "normal"
		elif (card.foil == "Etched"): foiling = "etch"
		else: foiling = "foil"
		exportFile.write(f"\"{card.name}\", {card.cn}, {card.set}, {foiling}, {card.quantity}\n")
	print("Done exporting")
	exportFile.close()

# Get prices
rowNumber = 0
addedCount = 0
done = 0
avgWaitTimes = [timeWait]

if (not args.stop): args.stop = len(cards)
print(f"Excel file set up. Retreiving prices from scryfall ({args.stop} cards), column {column}")
mm.check_and_reset_cache()
for card in cards:
	if (args.export_only and not args.validate): break

	# Time watching and early stopping
	if (done >= args.stop): break
	start = time.time()
	# Percent done
	perc = round((done / args.stop) * 100, 2)
	eta = np.average(avgWaitTimes) * (args.stop - done)
	eta = mm.convTime(eta)
	
	if (not args.validate or args.export_only):
		# Not validating
		""" Excel sheet: A - Card, B - Collector Number (CN), C - Foiling, D - Set, E: - Date* """
		try:
			singlePrice = mm.getPrice(card, args.use_cache)
    
		except mm.InvalidCardException as ICE:
			mm.log(f"WARNING: {ICE}", printMsg=True)
			continue
		except Exception as unknownError:
			mm.log(f"ERROR: Unknown error on {card}\n\t{unknownError}", close=True, printMsg=True)
		
		# Search for an already existing cell with card name, collector number, foiling, and set
		rowNumber = 1
		found = False
		while (sheet[f"A{rowNumber}"].value != None):
			excelCardInfo = [sheet[f"A{rowNumber}"].value, str(sheet[f"B{rowNumber}"].value), sheet[f"C{rowNumber}"].value, sheet[f"D{rowNumber}"].value]
			acessCardInfo = [card.name, str(card.cn), card.foil, card.set]
			compared = mm.getDifferences(excelCardInfo, acessCardInfo)
			if (mm.allTrue(compared)):
				sheet[f"{column}{rowNumber}"] = singlePrice
				sheet[f"{column}{rowNumber}"].number_format = '"$"#,##0.00_);("$"#,##0.00)'
				shortLine = f"\r\t{perc}%, eta = {eta}"
				line = f"\r\t{perc}% - Updated {card.name} ({card.cn} {card.set}, {card.foil}) for {singlePrice}, eta = {eta}"
				if (args.print): print(line + " " * len(line), flush = True, end = "")
				else: print(shortLine + " " * len(shortLine), flush = True, end = "")
				found = True
				break
			else: rowNumber += 1
		
		# Went through all cards in sheet and card was not found. Add at latest checked rowNumber
		if (not found):
			sheet[f"A{rowNumber}"] = card.name
			sheet[f"B{rowNumber}"] = card.cn
			sheet[f"C{rowNumber}"] = card.foil
			sheet[f"D{rowNumber}"] = card.set
			sheet[f"{column}{rowNumber}"] = singlePrice
			sheet[f"{column}{rowNumber}"].number_format = '"$"#,##0.00_);("$"#,##0.00)'
			shortLine = f"\r\t{perc}%, eta = {eta}"
			line = f"\r\t{perc}% - Added {card.name} ({card.cn} {card.set}, {card.foil}) for {singlePrice}, eta = {eta}"
			if (args.print): print(line + " " * len(line), flush = True, end = "")
			else: print(shortLine + " " * len(shortLine), flush = True, end = "")
			addedCount += 1
	
	elif (args.validate):
		same, trueName = mm.validate(card)
		line = f"Record shows '{card}' but got {trueName}"
		percStr = f"\t[{perc}%] "
		
		# Print short information
		if (not args.print): print("\r" + percStr, end = "", flush = True)
		# Print full information
		else:
			# Cards do not equal
			if (not same):
				fLine = percStr + line
				print(fLine, end = "\n", flush = True)
		
		if (not same): validationFile.write(line + "\n")
	
	rowNumber += 1
	done += 1
	# time.sleep(timeWait)
	end = time.time()
	avgWaitTimes.append(end - start)

if (not args.validate):
	print(f"Added {addedCount} new cards")
	print("All cards added and updated, closing files")

# Close everything
workbook.save(filename = args.name)
if (args.validate): validationFile.close()
print("All files closed and saved. You may exit this program and access the files now")
if (not args.close): input()
mm.log("INFO: Finished")