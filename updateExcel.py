from openpyxl import Workbook, load_workbook
import datetime

workbook = None
try:
	workbook = load_workbook(filename = "MagicPrices.xlsx")
except:
	workbook = Workbook()
	workbook.save(filename = "MagicPrices.xlsx")

sheet = workbook.active

print(sheet["B1"].value)

workbook.save(filename = "MagicPrices.xlsx")