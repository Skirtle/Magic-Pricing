import time, pyodbc, datetime
import magicPrices as mp
from openpyxl import Workbook, load_workbook

# Final variables
accessFilename = "Magic.accdb"
excelFilename = "MagicPrices.xlsx"
now = datetime.datetime.now()
timeWait = 0.2

# Connection to database
cnxn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=C:\Users\Dalton\Desktop\Magic-Pricing\\' + accessFilename + ";")
cursor = cnxn.cursor()
cnxn.setdecoding(pyodbc.SQL_CHAR, encoding='utf-8')
cnxn.setdecoding(pyodbc.SQL_WCHAR, encoding='utf-8')
cnxn.setencoding(encoding='utf-8')

# Open excel workbook
workbook = None
try:
	workbook = load_workbook(filename = excelFilename)
except:
	workbook = Workbook()
	workbook.save(filename = excelFilename)
sheet = workbook.active
sheet["A1"] = "Card"
sheet["B1"] = "CN"
sheet["C1"] = "Foiling"
sheet["D1"] = "Set"

# Get all cards
print("Accessing database for cards")
cursor.execute("select * from Cards")
rows = cursor.fetchall()
cards = [mp.Card(c[0], c[1], c[2], c[3]) for c in rows]
print("Cards collected. Setting up excel file for new entries")

# Get new column for prices
column = 1
while (sheet[f"{chr(ord('@') + column)}1"].value != None):
	# print(sheet[f"A{column}"].value)
	column += 1
column = chr(ord('@') + column)
sheet[f"{column}1"] = datetime.datetime(now.year, now.month, now.day)
sheet[f"{column}1"].number_format = "mm/dd/yyyy;@"

# Get prices
print("Excel file set up. Retreiving prices from scryfall")
price = 0
rowNumber = 0
for card in cards:
	singlePrice = mp.getPrice(card)
	
	# Find already existing cell with CARDNAME
	rowNumber = 1
	found = False
	while (sheet[f"A{rowNumber}"].value != None):
		# print(f"Checking card at A{rowNumber}, {sheet[f'A{rowNumber}'].value} == {card.name}")
		if (sheet[f"A{rowNumber}"].value == card.name and str(sheet[f"B{rowNumber}"].value) == str(card.cn) and sheet[f"C{rowNumber}"].value == card.foil and sheet[f"D{rowNumber}"].value == card.set):
			sheet[f"{column}{rowNumber}"] = singlePrice
			sheet[f"{column}{rowNumber}"].number_format = '"$"#,##0.00_);("$"#,##0.00)'
			# print(f"\t{card.name} found")
			found = True
			break
		else:
			rowNumber += 1
			
	if (not found):
		sheet[f"A{rowNumber}"] = card.name
		sheet[f"B{rowNumber}"] = card.cn
		sheet[f"C{rowNumber}"] = card.foil
		sheet[f"D{rowNumber}"] = card.set
		sheet[f"{column}{rowNumber}"] = singlePrice
		sheet[f"{column}{rowNumber}"].number_format = '"$"#,##0.00_);("$"#,##0.00)'
	rowNumber += 1
	print(f"\tAdded {card.name} ({card.cn} {card.set}, {card.foil}) for {singlePrice}")
	time.sleep(timeWait)
print("All cards added to excel, closing files.")

# Close everything
cursor.close()
cnxn.close()
workbook.save(filename = excelFilename)
print("All files closed and saved. You may exit this program and access the updated files now")